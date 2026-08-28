#!/usr/bin/env python3
"""Fold the office's "ALL CATEGORIES and Join Date" export into data/directory.json.

Felicia, Aug 28 2026: the ChamberWare export lists every category a company
carries (Category 1-3) plus its join date.  The site had only ONE category per
member, so a business that is both a caterer and an event venue could only ever
be found under one of them.  This merges the full set into `categories[]`, which
the directory filter and Wendy already search.

Two rules this script will not break:

  * Join dates already on the site stay put.  A blank is filled in from the
    export, but an existing date is never overwritten -- the export's own notes
    flag that duplicate company rows (franchises like Edward Jones, State Farm)
    fall back to "the latest available join date", which is a guess, and the
    office asked for their dates to remain in place.
  * The export carries member emails.  They are used in memory to disambiguate
    duplicate company names and are never written to the committed seed, which
    deliberately holds no member email addresses.

Usage:  python3 scripts/import-directory-categories.py <export.xlsx> [--dry-run]
Needs:  pip install openpyxl   (the .xlsx itself is gitignored PII -- keep it
        out of the repo and delete it when you are done.)
"""
import sys, os, re, json, difflib, collections

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl is required:  pip install openpyxl')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIRECTORY = os.path.join(ROOT, 'data', 'directory.json')
TAXONOMY = os.path.join(ROOT, 'data', 'category-groups.json')

# Legal-form noise and articles that differ between the export and the site
# ("Boulevard Banquet Hall, The" vs "The Boulevard Banquet Hall").
NOISE = re.compile(r'\b(the|inc|llc|llp|corp|corporation|co|company|of|apc|pc|dds|a)\b')


def norm(s):
    s = str(s or '').lower().replace('&', ' and ')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', NOISE.sub(' ', s)).strip()


def load_groups():
    """[(group, compiled matcher)] -- keywords match at a word start, so 'spa'
    does not match 'newspaper' and 'pet' does not match 'carpet'."""
    with open(TAXONOMY, encoding='utf-8') as fh:
        groups = json.load(fh)['groups']
    out = []
    for g in groups:
        pattern = '|'.join(r'\b' + re.escape(k) for k in g['keywords'])
        out.append((g['name'], re.compile(pattern)))
    return out


def group_of(member, groups):
    """Broad parent group, now decided by EVERY category a member carries."""
    hay = ' '.join(filter(None, [
        member.get('category', ''),
        ' '.join(member.get('categories') or []),
        member.get('typeOfBusiness', ''),
    ])).lower()
    if hay.strip():
        for name, matcher in groups:
            if matcher.search(hay):
                return name
    # Nothing matched: keep a group the office set by hand rather than
    # demoting a curated member to "Other".
    return member.get('group') or 'Other'


def read_export(path):
    """company key -> {'categories': [...], 'joinDate': 'YYYY-MM-DD' or None}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or '').strip().upper() for h in next(rows)]
    col = {name: header.index(name) for name in header}

    def cell(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = {}
    for row in rows:
        company = cell(row, 'COMPANY')
        if not company:
            continue
        rec = out.setdefault(norm(company), {'categories': [], 'joinDate': None, 'name': company})
        for c in ('CATEGORY 1', 'CATEGORY 2', 'CATEGORY 3'):
            v = cell(row, c)
            if v and str(v).strip() and str(v).strip() not in rec['categories']:
                rec['categories'].append(str(v).strip())
        jd, match = cell(row, 'JOINDATE'), str(cell(row, 'JOIN DATE MATCH') or '')
        if jd and rec['joinDate'] is None and match != 'No match':
            rec['joinDate'] = jd.strftime('%Y-%m-%d') if hasattr(jd, 'strftime') else str(jd)[:10]
    return out


def match_members(members, export):
    """Exact normalized name, then a conservative fuzzy/containment pass."""
    pairs, used = {}, set()
    for m in members:
        k = norm(m['name'])
        if k in export:
            pairs[m['id']] = k
            used.add(k)

    leftover_keys = [k for k in export if k not in used]
    for m in members:
        if m['id'] in pairs or not leftover_keys:
            continue
        k = norm(m['name'])
        close = difflib.get_close_matches(k, leftover_keys, n=1, cutoff=0.86)
        if close:
            pairs[m['id']] = close[0]
            leftover_keys.remove(close[0])
            continue
        tokens = set(k.split())
        if len(tokens) < 2:
            continue
        for other in leftover_keys:                     # one name fully contains the other
            ot = set(other.split())
            if len(ot) >= 2 and (tokens <= ot or ot <= tokens):
                pairs[m['id']] = other
                leftover_keys.remove(other)
                break
    return pairs


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    dry = '--dry-run' in sys.argv
    if not args:
        sys.exit(__doc__)
    export = read_export(args[0])
    groups = load_groups()
    with open(DIRECTORY, encoding='utf-8') as fh:
        doc = json.load(fh)
    members = doc['members']
    pairs = match_members(members, export)

    stats = collections.Counter()
    multi = 0
    for m in members:
        key = pairs.get(m['id'])
        before_group = m.get('group')
        if key:
            rec = export[key]
            cats = list(rec['categories'])
            primary = m.get('category')
            if primary and primary not in cats:         # never lose the category the site already showed
                cats.insert(0, primary)
            if not primary and cats:
                m['category'] = cats[0]
                stats['primary filled'] += 1
            if cats:
                if m.get('categories') != cats:
                    stats['categories set'] += 1
                m['categories'] = cats
                if len(cats) > 1:
                    multi += 1
            if rec['joinDate']:
                if not m.get('joinDate'):
                    m['joinDate'] = rec['joinDate']
                    stats['join date filled'] += 1
                elif m['joinDate'] != rec['joinDate']:
                    stats['join date kept (export differs)'] += 1
        else:
            stats['no export row'] += 1
        m['group'] = group_of(m, groups)
        if m['group'] != before_group:
            stats['group changed'] += 1

    doc['_meta'] = {**doc.get('_meta', {}),
                    'categoriesImportedAt': '2026-08-28',
                    'categoriesSource': 'office export: ALL CATEGORIES and Join Date'}

    print(f'members            : {len(members)}')
    print(f'matched to export  : {len(pairs)}')
    print(f'carry >1 category  : {multi}')
    for k, v in sorted(stats.items()):
        print(f'  {k:32s} {v}')
    dist = collections.Counter(m.get('group') for m in members)
    print('\ngroups (' + str(len(dist)) + '):')
    for g, n in dist.most_common():
        print(f'  {n:4d}  {g}')

    if dry:
        print('\n--dry-run: nothing written')
        return
    with open(DIRECTORY, 'w', encoding='utf-8') as fh:
        json.dump(doc, fh, indent=2, ensure_ascii=False)
        fh.write('\n')
    print('\nwrote data/directory.json')


if __name__ == '__main__':
    main()
